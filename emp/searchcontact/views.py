
from django.shortcuts import render

from openpyxl import load_workbook
from .forms import searchform
# Create your views here.
def search(request):
    form=searchform()
    items=[]
    if request.method=='POST':
        form=searchform(request.POST)
        if form.is_valid():
            cd=form.cleaned_data
            wp=load_workbook(r'C:\Users\ayaga\project\emp\names.xlsx')
            ws = wp['emp']
            noofrows=int(ws['F1'].value)
            for i in range(2,noofrows+1):
                if cd['name']==ws['A' + str(i)].value:
                    item=[]
                    item.append(ws['A' + str(i)].value)
                    item.append(ws['B' + str(i)].value)
                    item.append(ws['C' + str(i)].value)
                    item.append(ws['D' + str(i)].value)
                    items.append(item)

    return render(request,'page/search.html',{'items':items,'form':form})