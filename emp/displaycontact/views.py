from django.shortcuts import render

# Create your views here.
from openpyxl import load_workbook

def disp(request):
    wp=load_workbook(r'C:\Users\ayaga\project\emp\names.xlsx')
    ws = wp['emp']
    noofrows=int(ws['F1'].value)
    items=[]
    for i in range(2,noofrows+1):
        item=[]
        item.append(ws['A' + str(i)].value)
        item.append(ws['B' + str(i)].value)
        item.append(ws['C' + str(i)].value)
        item.append(ws['D' + str(i)].value)
        items.append(item)

    return render(request,'page/displaydata.html',{'items':items})
