import string
from django import forms
from django.shortcuts import render
from .forms import RegEmpForm
from openpyxl import load_workbook

def regEmp(request):

    m = ''

    form = RegEmpForm()
    if request.method == 'POST':
        form = RegEmpForm(request.POST)
        if form.is_valid():
            cd =form.cleaned_data
            wb = load_workbook(r'C:\Users\ayaga\project\newsproject\news.xlsx')
            ws = wb['news']

            i= str(string(ws['F1'].value)+1)
            
            ws['A' + i] = cd['name']
           

            ws['F1'] = i

            wb.save(r'C:\Users\ayaga\project\newsproject\news.xlsx')
            m = 'Saved Successfully'

            form = RegEmpForm()

    return render(request, 'page/newemp.html', {'form': form , 'output': m })

